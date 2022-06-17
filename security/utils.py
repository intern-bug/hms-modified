from institute.models import Block
from security.models import OutingInOutTimes
from django.utils import timezone
from openpyxl import Workbook
from openpyxl import styles

class OutingBookGenerator:
    def __init__(self, block_id, year_month_day):
        self.block_id = block_id
        self.year_month_day = year_month_day
    
    def generate_workbook(self):
        workbook  = Workbook()
        workbook.remove(workbook.active)

        if self.block_id == 'all':
            for block in Block.objects.all():
                SheetGenerator = BlockOutingSheetGenerator(block.id, self.year_month_day)
                SheetGenerator.generate_block_sheet(workbook)
        else:
            SheetGenerator = BlockOutingSheetGenerator(self.block_id, self.year_month_day)
            SheetGenerator.generate_block_sheet(workbook)

        return workbook


class BlockOutingSheetGenerator:
    def __init__(self, block_id, year_month_day):
        self.block = Block.objects.get(id = block_id)
        if year_month_day == 'all':
            self.day = "all"
            self.month = 'all'
            self.year = 'all'
            self.outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).order_by('outing__fromDate')
        else:
            self.year, self.month, self.day = [int(x) for x in year_month_day.split("-")]
            if self.year and self.month and self.day:
                from_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__fromDate__date=year_month_day)
                to_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__toDate__date=year_month_day)
            elif self.year and self.month:
                from_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__fromDate__year=self.year).filter(outing__fromDate__month=self.month)
                to_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__toDate__year=self.year).filter(outing__toDate__month=self.month)
            elif self.year:
                from_outing_list = OutingInOutTimes.objects.filter(outing__fromDate__year=self.year)
                to_outing_list = OutingInOutTimes.objects.filter(outing__toDate__year=self.year)
            outing_list = (from_outing_list|to_outing_list).distinct()
            outing_list = outing_list.order_by('outing__fromDate')
            self.outing_list=outing_list
        
    def generate_block_sheet(self, workbook):
        worksheet = workbook.create_sheet(title = "{}".format(self.block.name))

        headers = ['Regd. No.', 'Name', 'Year', 'Specialization', 'Outing Mode', 'From Date', 'Out Time', 'To Date', 'In Time']
        row_num = 1

        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold = True)
            cell.value = column_title

        for outingInOut in self.outing_list:
            row_num += 1
            row_data = [outingInOut.outing.student.regd_no, 
                        outingInOut.outing.student.name,
                        outingInOut.outing.student.year,
                        outingInOut.outing.student.branch,
                        outingInOut.outing.type,
                        outingInOut.outing.fromDate.strftime("%d-%m-%Y, %H:%M:%S"),
                        outingInOut.outTime.strftime("%d-%m-%Y, %H:%M:%S"),
                        outingInOut.outing.toDate.strftime("%d-%m-%Y, %H:%M:%S"),
                        ]
            if outingInOut.inTime:
                row_data.append(outingInOut.inTime.strftime("%d-%m-%Y, %H:%M:%S"))
            else:
                row_data.append('')
            row_color = self.get_cell_color(outingInOut)
            for col_num, cell_value in enumerate(row_data,1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = styles.Font(color=row_color)
                
    def get_cell_color(self,outingInOutObj):
        if outingInOutObj.outing.type == 'Local':
            if outingInOutObj.student.gender == 'Male' and outingInOutObj.inTime != None:
                if ((outingInOutObj.inTime - outingInOutObj.outing.toTime).total_seconds()/60) > 15.0:
                    return 'DC3545'
                if (outingInOutObj.inTime.hour*100 + outingInOutObj.inTime.minute) > 2115 :
                    return 'DC3545'
                else:
                    return '28A745'
            elif outingInOutObj.student.gender == 'Female' and outingInOutObj.inTime != None:
                if ((outingInOutObj.inTime - outingInOutObj.outing.toTime).total_seconds()/60) > 15.0:
                    return 'DC3545'
                if (outingInOutObj.inTime.hour*100 + outingInOutObj.inTime.minute) > 2045 :
                    return 'DC3545'
                else:
                    return '28A745'
            elif outingInOutObj.inTime == None and ((timezone.now() - outingInOutObj.outing.toDate).total_seconds()/60) < 15.0:
                return '28A745'
            elif outingInOutObj.inTime == None and ((timezone.now() - outingInOutObj.outing.toDate).total_seconds()/60) > 15.0:
                return 'DC3545'
        else:
            if outingInOutObj.inTime != None and ((outingInOutObj.inTime - outingInOutObj.outing.toDate).total_seconds()/3600) > 1.0:
                return 'DC3545'
            elif outingInOutObj.inTime == None and (((timezone.now() - outingInOutObj.outing.toDate).total_seconds()/3600) > 1.0):
                return 'DC3545'
            else:
                return '28A745'
