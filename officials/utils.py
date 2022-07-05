from institute.models import Block
from students.models import Attendance
from security.models import OutingInOutTimes
from django.utils import timezone
from openpyxl import Workbook
from openpyxl import styles

class AttendanceBookGenerator:
    def __init__(self, block_id, year_month):
        self.block_id = block_id
        self.year_month = year_month
    
    def generate_workbook(self):
        workbook  = Workbook()
        workbook.remove(workbook.active)

        if self.block_id == 'all':
            for block in Block.objects.all():
                SheetGenerator = BlockAttendanceSheetGenerator(block.id, self.year_month)
                SheetGenerator.generate_block_sheet(workbook)
        elif self.block_id == 'boys':
            for block in Block.objects.filter(gender='Male'):
                SheetGenerator = BlockAttendanceSheetGenerator(block.id, self.year_month)
                SheetGenerator.generate_block_sheet(workbook)
        elif self.block_id == 'girls':
            for block in Block.objects.filter(gender='Female'):
                SheetGenerator = BlockAttendanceSheetGenerator(block.id, self.year_month)
                SheetGenerator.generate_block_sheet(workbook)
        else:
            SheetGenerator = BlockAttendanceSheetGenerator(self.block_id, self.year_month)
            SheetGenerator.generate_block_sheet(workbook)

        return workbook


class BlockAttendanceSheetGenerator:
    def __init__(self, block_id, year_month):
        self.block = Block.objects.get(id = block_id)
        self.attendance_list = Attendance.objects.filter(student__in = self.block.roomdetail_set.all().values_list('student', flat=True))
        if year_month == 'all':
            self.month = 'all'
            self.year = 'all'
        else:
            self.year, self.month = [int(x) for x in year_month.split("-")]

    def generate_block_sheet(self, workbook):
        worksheet = workbook.create_sheet(title = "{}".format(self.block.name))

        self.generate_dates()
        headers = ['Regd. No.', 'Name'] + [date.strftime("%d/%m/%y") for date in self.generated_dates]
        row_num = 1

        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold = True)
            cell.value = column_title

        for attendance in self.attendance_list:
            row_num += 1
            row_data = [attendance.student.regd_no, attendance.student.name] + self.get_student_attendance(attendance)

            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if cell_value == 'P': cell.font = styles.Font(color='28A745')
                elif cell_value == 'A': cell.font = styles.Font(bold=True, color='DC3545')


    def get_student_attendance(self, attendance):
        present_absent_list = []
        present_dates = None
        absent_dates = None
        if attendance.present_dates:
            present_dates = attendance.present_dates.split(',')
            present_dates = [present_date.split('@')[0] for present_date in present_dates]
        if attendance.absent_dates:
            absent_dates = attendance.absent_dates.split(',')
            absent_dates = [absent_date.split('@')[0] for absent_date in absent_dates]
        for date in self.generated_dates:
            date_formatted = date.strftime("%Y-%m-%d")
            if present_dates and  date_formatted in present_dates:
                present_absent_list.append('P')
            elif absent_dates and date_formatted in absent_dates:
                present_absent_list.append('A')
            else:
                present_absent_list.append('-')
        return present_absent_list

    def generate_dates(self):
        if self.year == 'all' or self.month == 'all':
            date_set = self.get_marked_dates()
            month_set = self.get_month_year_set(date_set)
            generated_dates = []

            for item in month_set:
                dates_of_month = self.get_dates_of_month(item[0], item[1])
                generated_dates += dates_of_month

            self.generated_dates = sorted(generated_dates)
        
        else:
            self.generated_dates = self.get_dates_of_month(self.month, self.year)

    def get_dates_of_month(self, month, year):
            day = timezone.timedelta(days=1)
            start_date = timezone.datetime(year = year, month = month, day = 1)
            dates_of_month = []
            d = start_date
            while d.month == month:
                dates_of_month.append(d)
                d += day

            return sorted(dates_of_month)

    def get_marked_dates(self):
        date_set = set()
        for attendance in self.attendance_list:
            if attendance.present_dates:
                present_dates = attendance.present_dates.split(',')
                present_dates = [present_date.split('@')[0] for present_date in present_dates]
                date_set |= set(present_dates)
            if attendance.absent_dates:
                absent_dates = attendance.absent_dates.split(',')
                absent_dates = [absent_date.split('@')[0] for absent_date in absent_dates]
                date_set |= set(absent_dates)
        date_set.discard('')

        date_format = "%Y-%m-%d"
        date_set = set([timezone.datetime.strptime(date, date_format) for date in date_set])
        self.marked_dates = date_set
        return self.marked_dates

    # Set of tuples (month, year) that have attendance marked
    def get_month_year_set(self, date_set):
        month_year_set = set([(date.month, date.year) for date in date_set])
        return month_year_set

class OutingBookGenerator:
    def __init__(self, block_id, year_month_day):
        self.block_id = block_id
        self.year_month_day = year_month_day
    
    def generate_workbook(self):
        workbook  = Workbook()
        workbook.remove(workbook.active)

        if self.block_id == 'all':
            for block in Block.objects.all():
                SheetGenerator = BlockOutingSheetGenerator(block.id, self.year_month_day)
                SheetGenerator.generate_block_sheet(workbook)
        elif self.block_id == 'boys':
            for block in Block.objects.filter(gender='Male'):
                SheetGenerator = BlockOutingSheetGenerator(block.id, self.year_month_day)
                SheetGenerator.generate_block_sheet(workbook)
        elif self.block_id == 'girls':
            for block in Block.objects.filter(gender='Female'):
                SheetGenerator = BlockOutingSheetGenerator(block.id, self.year_month_day)
                SheetGenerator.generate_block_sheet(workbook)
        else:
            SheetGenerator = BlockOutingSheetGenerator(self.block_id, self.year_month_day)
            SheetGenerator.generate_block_sheet(workbook)

        return workbook


class BlockOutingSheetGenerator:
    def __init__(self, block_id, year_month_day):
        self.block = Block.objects.get(id = block_id)
        if year_month_day == 'all':
            self.day = "all"
            self.month = 'all'
            self.year = 'all'
            self.outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).order_by('outing__fromDate')
        else:
            self.year, self.month, self.day = [int(x) for x in year_month_day.split("-")]
            if self.year and self.month and self.day:
                from_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__fromDate__date=year_month_day)
                to_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__toDate__date=year_month_day)
            elif self.year and self.month:
                from_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__fromDate__year=self.year).filter(outing__fromDate__month=self.month)
                to_outing_list = OutingInOutTimes.objects.filter(outing__student__in=self.block.students()).\
                    filter(outing__toDate__year=self.year).filter(outing__toDate__month=self.month)
            elif self.year:
                from_outing_list = OutingInOutTimes.objects.filter(outing__fromDate__year=self.year)
                to_outing_list = OutingInOutTimes.objects.filter(outing__toDate__year=self.year)
            outing_list = (from_outing_list|to_outing_list).distinct()
            outing_list = outing_list.order_by('outing__fromDate')
            self.outing_list=outing_list
        
    def generate_block_sheet(self, workbook):
        worksheet = workbook.create_sheet(title = "{}".format(self.block.name))

        headers = ['Regd. No.', 'Name', 'Year', 'Specialization', 'Outing Mode', 'From Date', 'Out Time', 'To Date', 'In Time']
        row_num = 1

        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold = True)
            cell.value = column_title

        for outingInOut in self.outing_list:
            row_num += 1
            row_data = [outingInOut.outing.student.regd_no, 
                        outingInOut.outing.student.name,
                        outingInOut.outing.student.year,
                        outingInOut.outing.student.branch,
                        outingInOut.outing.type,
                        outingInOut.outing.fromDate.strftime("%d-%m-%Y, %H:%M:%S"),
                        outingInOut.outTime.strftime("%d-%m-%Y, %H:%M:%S"),
                        ]
            if outingInOut.outing.type == 'Vacation':
                row_data.append('')
            else:
                row_data.append(outingInOut.outing.toDate.strftime("%d-%m-%Y, %H:%M:%S"))
            if outingInOut.inTime and outingInOut.outing.type != 'Vacation':
                row_data.append(outingInOut.inTime.strftime("%d-%m-%Y, %H:%M:%S"))
            else:
                row_data.append('')
            row_color = self.get_cell_color(outingInOut)
            for col_num, cell_value in enumerate(row_data,1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = styles.Font(color=row_color)
                
    def get_cell_color(self,outingInOutObj):
        if outingInOutObj.outing.type == 'Local':
            if outingInOutObj.outing.student.gender == 'Male' and outingInOutObj.inTime != None:
                if (outingInOutObj.inTime.hour*100 + outingInOutObj.inTime.minute) > 2115 :
                    return 'DC3545'
                else:
                    return '28A745'
            elif outingInOutObj.outing.student.gender == 'Female' and outingInOutObj.inTime != None:
                if (outingInOutObj.inTime.hour*100 + outingInOutObj.inTime.minute) > 2045 :
                    return 'DC3545'
                else:
                    return '28A745'
            elif outingInOutObj.inTime == None and ((timezone.now() - outingInOutObj.outing.toDate).total_seconds()/60) < 15.0:
                return '28A745'
            elif outingInOutObj.inTime == None and ((timezone.now() - outingInOutObj.outing.toDate).total_seconds()/60) > 15.0:
                return 'DC3545'
        else:
            if outingInOutObj.inTime != None and ((outingInOutObj.inTime - outingInOutObj.outing.toDate).total_seconds()/3600) > 1.0:
                return 'DC3545'
            elif outingInOutObj.inTime == None and (((timezone.now() - outingInOutObj.outing.toDate).total_seconds()/3600) > 1.0):
                return 'DC3545'
            else:
                return '28A745'

class MessReportBookGenerator:
    def __init__(self, rebate_list):
        self.rebate_list = rebate_list
    
    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_sheet(workbook=workbook)

        return workbook
    
    def generate_sheet(self, workbook):
        file = str(timezone.localtime().strftime("%d-%m-%Y_%H-%M-%S"))
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        headers = ['Regd. No.', 'Name', 'From', 'To', 'No. of Days', 'No. of rebate days', 'No. of effective days']
        row_num = 1
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font =styles.Font(bold = True)
            cell.value = column_title
        for rebate in self.rebate_list:
            row_num += 1
            row_data = [
                rebate['outing__student__regd_no'],
                rebate['outing__student__name'],
                rebate['from_date'],
                rebate['to_date'],
                rebate['total_days'],
                rebate['no_of_days'],
                rebate['effective_days']
            ]
            for col_num, cell_value in enumerate(row_data,1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


